"""
Excel Parser
Парсинг файлов нагрузки из Excel
"""

import openpyxl
import logging
from typing import List, Dict, Any, Optional, Tuple
import io
import re
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelParser:
    """Парсер Excel файлов с нагрузкой"""
    
    @staticmethod
    def parse_course_loads(
        file_data: bytes,
        semester: int = None,
        academic_year: str = None
    ) -> tuple[List[Dict[str, Any]], List[str]]:
        """
        Парсить нагрузку из Excel
        
        Поддерживает два формата:
        1. Простой формат (10 колонок по порядку):
           - Дисциплина, Код дисциплины, Преподаватель, ID преподавателя, 
             Приоритет, Группа, ID группы, Размер группы, Тип занятия, Часы
        
        2. Расширенный формат (по названиям колонок):
           - Дисциплина -> discipline_name
           - Блок -> discipline_code (извлекается код)
           - Преподаватель -> teacher_name
           - Группа -> group_name
           - Нагрузка -> lesson_type
           - Часы -> hours_per_semester
           - Количество контингента -> group_size
           - Семестр -> извлекается номер семестра ИЗ КАЖДОЙ СТРОКИ
        
        КРИТИЧНО:
        - Учебный год парсится из файла (из первых строк, паттерн "2025-2026 уч. год")
        - Семестр парсится из каждой строки (колонка "Семестр")
        - Если не найдены в файле, используются параметры semester и academic_year
        
        Args:
            file_data: Байты Excel файла
            semester: Номер семестра (используется если не найден в файле)
            academic_year: Учебный год (используется если не найден в файле)
        
        Returns:
            (список_нагрузок, список_ошибок)
        """
        course_loads = []
        errors = []
        
        try:
            # Открыть Excel
            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
            ws = wb.active
            
            # Найти строку с заголовками (не всегда первая строка!)
            header_row_idx = None
            for row_idx in range(1, min(20, ws.max_row + 1)):  # Проверяем первые 20 строк
                row_values = [
                    str(cell.value).strip().lower() if cell.value else ""
                    for cell in ws[row_idx]
                ]
                
                # Проверяем наличие ключевых слов заголовков
                has_discipline = any('дисциплина' in val for val in row_values)
                has_teacher = any('преподаватель' in val for val in row_values)
                has_hours = any('час' in val for val in row_values)
                has_group = any('группа' in val for val in row_values)
                has_load = any('нагрузка' in val for val in row_values)
                
                # Если нашли несколько ключевых колонок - это заголовки
                found_count = sum([
                    has_discipline, has_teacher, has_hours, has_group, has_load
                ])
                
                if found_count >= 3:  # Нашли минимум 3 ключевые колонки
                    header_row_idx = row_idx
                    logger.info(
                        f"Found header row at index {row_idx} "
                        f"(found {found_count} key columns)"
                    )
                    break
            
            if header_row_idx is None:
                # Если не нашли, используем первую строку
                header_row_idx = 1
                logger.warning(
                    "Could not find header row, using row 1 as default"
                )
            
            # Получить заголовки из найденной строки
            headers = []
            for cell in ws[header_row_idx]:
                headers.append(str(cell.value).strip() if cell.value else "")
            
            logger.info(f"Found headers (row {header_row_idx}): {headers}")
            
            # Парсить учебный год из файла (из первых строк или заголовков)
            parsed_academic_year = academic_year
            if not parsed_academic_year:
                parsed_academic_year = ExcelParser._extract_academic_year(ws, header_row_idx)
                if parsed_academic_year:
                    logger.info(f"Extracted academic year from file: {parsed_academic_year}")
                else:
                    logger.warning("Could not extract academic year from file")
            
            # Определить формат файла и найти индексы колонок
            column_map = {}
            
            # Попытка найти колонки по названиям (расширенный формат)
            header_lower = [h.lower() for h in headers]
            
            # Маппинг названий колонок (более гибкий поиск)
            column_mappings = {
                'discipline_name': ['дисциплина', 'название дисциплины'],
                'discipline_code': ['блок', 'код дисциплины', 'код'],
                'teacher_name': ['преподаватель', 'фио преподавателя', 'фио'],
                'group_name': ['группа', 'название группы'],
                'lesson_type': ['нагрузка', 'тип занятия', 'вид занятия'],
                'hours': [
                    'часы', 'час', 'часов', 'часов в семестр',
                    'количество часов', 'астрономических часов'
                ],
                'group_size': [
                    'количество контингента',  # Приоритет: точное совпадение
                    'размер группы',
                    'количество'  # Убрали 'контингент' - слишком общее
                ],
                'semester_text': ['семестр', 'сем'],
                'contingent': ['контингент']
            }
            
            # Более точный поиск колонок
            # Сначала сортируем названия по длине (от длинных к коротким)
            # чтобы более точные совпадения имели приоритет
            for field, possible_names in column_mappings.items():
                # Сортируем по длине (от длинных к коротким)
                sorted_names = sorted(possible_names, key=len, reverse=True)
                
                for idx, header in enumerate(header_lower):
                    if field in column_map:
                        break
                    
                    header_clean = header.strip()
                    
                    for name in sorted_names:
                        # Точное совпадение (без учета регистра)
                        if header_clean == name:
                            column_map[field] = idx
                            logger.debug(
                                f"Found column '{field}' at index {idx}: "
                                f"'{headers[idx]}' (exact match: '{name}')"
                            )
                            break
                        # Вхождение слова (но не как часть другого слова)
                        elif name in header_clean and len(header_clean) > 0:
                            # Проверяем, что это не часть другого слова
                            # (например, "часы" не должно совпадать с "часовой")
                            if (header_clean.startswith(name) or
                                header_clean.endswith(name) or
                                f' {name} ' in f' {header_clean} '):
                                column_map[field] = idx
                                logger.debug(
                                    f"Found column '{field}' at index {idx}: "
                                    f"'{headers[idx]}' (partial match: '{name}')"
                                )
                                break
            
            # Если нашли колонки по названиям - используем расширенный формат
            # Также используем расширенный формат, если файл имеет много колонок (>10)
            has_required_columns = (
                'discipline_name' in column_map and 'teacher_name' in column_map
            )
            # КРИТИЧНО: если файл имеет >10 колонок, всегда используем расширенный формат
            use_named_columns = has_required_columns or len(headers) > 10
            
            # Если используем расширенный формат, но не нашли обязательные колонки
            if use_named_columns and not has_required_columns:
                logger.warning(
                    "Using extended format (many columns), but some required "
                    "columns not found. Will try to parse anyway."
                )
            
            if use_named_columns:
                logger.info(
                    f"Using named column mapping (extended format). "
                    f"Found columns: {column_map}"
                )
                logger.info(f"Total headers: {len(headers)}")
                logger.info(f"Header names: {headers[:10]}...")  # Первые 10
                
                # Проверить, что обязательные колонки найдены
                if 'hours' not in column_map:
                    msg = (
                        "Column 'Часы' not found in file. "
                        "Please check column names."
                    )
                    errors.append(msg)
                    logger.error(f"Column 'Часы' not found! Headers: {headers}")
                    # Попробовать найти колонку вручную
                    for idx, h in enumerate(headers):
                        if 'час' in h.lower():
                            logger.warning(
                                f"Found potential 'hours' column at {idx}: '{h}'"
                            )
            else:
                logger.info("Using positional column mapping (simple format)")
            
            # Парсить строки (начинаем со строки после заголовков)
            data_start_row = header_row_idx + 1
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=data_start_row, values_only=True),
                start=data_start_row
            ):
                try:
                    # Проверить, что строка не пустая
                    if not row or not any(row):
                        continue
                    
                    if use_named_columns:
                        # Расширенный формат - извлечение по названиям колонок
                        discipline_name = str(row[column_map.get('discipline_name', 0)] or '').strip()
                        if not discipline_name:
                            continue
                        
                        # Код дисциплины из блока (например, "Б1.В.01.ДВ.02.01")
                        discipline_code = ''
                        if 'discipline_code' in column_map:
                            block = str(row[column_map['discipline_code']] or '').strip()
                            discipline_code = block if block else ''
                        
                        teacher_name = str(row[column_map.get('teacher_name', 0)] or '').strip()
                        # Пропустить строки с "Вакансия"
                        if not teacher_name or teacher_name.lower() in ['вакансия', 'vacancy', '']:
                            continue
                        
                        group_name = str(row[column_map.get('group_name', 0)] or '').strip()
                        if not group_name:
                            continue
                        
                        # Тип занятия из колонки "Нагрузка"
                        lesson_type_raw = str(row[column_map.get('lesson_type', 0)] or '').strip()
                        lesson_type = ExcelParser._normalize_lesson_type(lesson_type_raw)
                        
                        # Часы - КРИТИЧНО: проверить, что колонка найдена
                        if 'hours' not in column_map:
                            # Если это расширенный формат, но колонка не найдена,
                            # это критическая ошибка
                            if row_idx == 2:  # Только для первой строки данных
                                errors.append(
                                    "Column 'Часы' not found in file. "
                                    "Cannot parse without hours column."
                                )
                                logger.error(
                                    "Column 'Часы' not found! "
                                    f"Available headers: {headers}"
                                )
                            continue
                        
                        hours_raw = row[column_map['hours']]
                        try:
                            hours_per_semester = (
                                float(hours_raw) if hours_raw else 0
                            )
                        except (ValueError, TypeError) as e:
                            errors.append(
                                f"Row {row_idx}: Invalid hours value "
                                f"'{hours_raw}': {e}"
                            )
                            continue
                        
                        if hours_per_semester <= 0:
                            errors.append(
                                f"Row {row_idx}: Hours must be positive, "
                                f"got {hours_per_semester}"
                            )
                            continue
                        
                        # Размер группы
                        group_size = None
                        if 'group_size' in column_map:
                            size_raw = row[column_map['group_size']]
                            try:
                                group_size = int(size_raw) if size_raw else None
                            except (ValueError, TypeError):
                                pass
                        
                        # Извлечь номер семестра из текста (если есть колонка "Семестр")
                        # КРИТИЧНО: семестр парсится из каждой строки!
                        file_semester = semester  # Дефолт, если не найден
                        semester_column_value = None
                        if 'semester_text' in column_map:
                            semester_text = str(
                                row[column_map['semester_text']] or ''
                            ).strip()
                            semester_column_value = semester_text  # Сохраняем для автоопределения
                            if semester_text:
                                extracted_semester = (
                                    ExcelParser._extract_semester_number(
                                        semester_text
                                    )
                                )
                                if extracted_semester:
                                    file_semester = extracted_semester
                        
                        # Извлечь информацию о контингенте (если есть колонка "Контингент")
                        contingent_info = None
                        if 'contingent' in column_map:
                            contingent_info = str(
                                row[column_map['contingent']] or ''
                            ).strip()
                        
                        # ID преподавателя и группы нужно будет найти в БД позже
                        # Используем NULL вместо 0, чтобы не нарушать foreign key constraints
                        teacher_id = None
                        teacher_priority = 2  # Дефолтный приоритет
                        group_id = None
                        
                    else:
                        # Простой формат - по позициям
                        if len(row) < 10:
                            errors.append(f"Row {row_idx}: Not enough columns (expected 10, got {len(row)})")
                            continue
                        
                        discipline_name = str(row[0] or '').strip()
                        if not discipline_name:
                            continue
                        
                        discipline_code = str(row[1] or '').strip()
                        teacher_name = str(row[2] or '').strip()
                        if not teacher_name or teacher_name.lower() in ['вакансия', 'vacancy']:
                            continue
                        
                        try:
                            teacher_id = int(row[3]) if row[3] else None
                            teacher_priority = int(row[4]) if row[4] else 2
                            group_name = str(row[5] or '').strip()
                            group_id = int(row[6]) if row[6] else None
                            group_size = int(row[7]) if row[7] else None
                            lesson_type = str(row[8] or '').strip()
                            hours_per_semester = float(row[9]) if row[9] else 0
                        except (ValueError, TypeError) as e:
                            errors.append(f"Row {row_idx}: Invalid data type - {e}")
                            continue
                        
                        file_semester = semester
                        semester_column_value = None
                        contingent_info = None
                    
                    # Вычислить пары в неделю
                    # Формула: часы / (1.5 часа * 16 недель) = часы / 24
                    # 1 пара = 2 академических часа = 90 минут =
                    # 1.5 астрономических часа
                    if hours_per_semester > 0:
                        lessons_per_week = max(1, round(hours_per_semester / 24))
                    else:
                        lessons_per_week = 1
                    
                    course_loads.append({
                        'discipline_name': discipline_name,
                        'discipline_code': discipline_code,
                        'teacher_id': teacher_id,
                        'teacher_name': teacher_name,
                        'teacher_priority': teacher_priority,
                        'group_id': group_id,
                        'group_name': group_name,
                        'group_size': group_size,
                        'lesson_type': lesson_type,
                        'hours_per_semester': hours_per_semester,
                        'lessons_per_week': lessons_per_week,
                        'semester': file_semester,
                        'academic_year': parsed_academic_year or academic_year,
                        # Дополнительные поля для автоопределения семестра
                        'semester_column': semester_column_value if use_named_columns else None,
                        'contingent_info': contingent_info if use_named_columns else None,
                        # source_file будет добавлен в parse_multiple_files
                        'source_file': ''
                    })
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {e}")
                    logger.error(f"Error parsing row {row_idx}: {e}", exc_info=True)
            
            wb.close()
            
            logger.info(f"Parsed {len(course_loads)} course loads, {len(errors)} errors")
            
            return course_loads, errors
            
        except Exception as e:
            logger.error(f"Failed to parse Excel: {e}", exc_info=True)
            errors.append(f"Parse error: {e}")
            return [], errors
    
    @staticmethod
    def _normalize_lesson_type(lesson_type: str) -> str:
        """
        Нормализовать тип занятия
        
        Маппит различные варианты названий на стандартные типы,
        но если тип не распознан, возвращает исходное значение
        (так как constraint теперь разрешает любые непустые значения)
        """
        if not lesson_type:
            return lesson_type
        
        lesson_type_lower = lesson_type.lower().strip()
        
        # Маппинг на стандартные типы (для удобства)
        if 'лекц' in lesson_type_lower or 'lecture' in lesson_type_lower:
            return 'Лекция'
        elif 'практик' in lesson_type_lower or 'practice' in lesson_type_lower:
            # "Практические занятия" -> "Практика"
            if 'практические занятия' in lesson_type_lower:
                return 'Практика'
            return 'Практика'
        elif 'лаборатор' in lesson_type_lower or 'lab' in lesson_type_lower:
            return 'Лабораторная'
        elif 'семинар' in lesson_type_lower or 'seminar' in lesson_type_lower:
            return 'Семинар'
        elif 'консультац' in lesson_type_lower or 'consultation' in lesson_type_lower:
            return 'Консультация'
        elif 'контрольн' in lesson_type_lower:
            return 'Контрольная работа'
        elif 'экзамен' in lesson_type_lower or 'exam' in lesson_type_lower:
            return 'Экзамен'
        elif 'зачет' in lesson_type_lower or 'зачёт' in lesson_type_lower:
            return 'Зачет'
        else:
            # Возвращаем исходное значение - constraint теперь разрешает любые типы
            return lesson_type.strip()
    
    @staticmethod
    def _extract_academic_year(ws, header_row_idx: int) -> str:
        """
        Извлечь учебный год из файла
        Ищет в первых строках паттерн типа "2025-2026 уч. год" или "2025/2026"
        """
        # Ищем в первых строках до заголовков
        for row_idx in range(1, min(header_row_idx + 1, 10)):
            row = ws[row_idx]
            for cell in row:
                if not cell.value:
                    continue
                cell_text = str(cell.value).strip()
                
                # Паттерн: "2025-2026 уч. год" или "2025/2026" или "2025-2026"
                # Ищем формат YYYY-YYYY или YYYY/YYYY
                year_pattern = r'(\d{4})[-\/](\d{4})'
                match = re.search(year_pattern, cell_text)
                if match:
                    year1, year2 = match.groups()
                    # Проверяем, что второй год = первый + 1
                    if int(year2) == int(year1) + 1:
                        academic_year = f"{year1}/{year2}"
                        logger.info(
                            f"Found academic year in row {row_idx}: "
                            f"'{cell_text}' -> {academic_year}"
                        )
                        return academic_year
        
        return None
    
    @staticmethod
    def _extract_semester_number(semester_text: str) -> int:
        """Извлечь номер семестра из текста (например, 'Третий семестр' -> 3)"""
        if not semester_text:
            return None
        
        semester_text_lower = semester_text.lower()
        
        # Словарь для русских числительных (расширенный до 12)
        russian_numbers = {
            'первый': 1, 'второй': 2, 'третий': 3, 'четвертый': 4,
            'пятый': 5, 'шестой': 6, 'седьмой': 7, 'восьмой': 8,
            'девятый': 9, 'десятый': 10, 'одиннадцатый': 11, 'двенадцатый': 12
        }
        
        for word, num in russian_numbers.items():
            if word in semester_text_lower:
                return num
        
        # Попытка извлечь число напрямую
        numbers = re.findall(r'\d+', semester_text)
        if numbers:
            return int(numbers[0])
        
        return None
    
    @staticmethod
    def extract_admission_year(group_name: str) -> Optional[int]:
        """
        Извлекает год поступления из названия группы.
        
        Args:
            group_name: Название группы, например "Б9124-01.03.02сп"
            
        Returns:
            Год поступления (2021) или None
            
        Examples:
            >>> extract_admission_year("Б9124-01.03.02сп")
            2021
            >>> extract_admission_year("Б9223-02.03.01")
            2022
        """
        if not group_name:
            return None
        
        # Извлекаем 4-значный код после буквы (Б9124, М9224)
        match = re.search(r'[А-Я](\d{4})', group_name)
        if not match:
            return None
        
        year_code = match.group(1)[:2]  # Берем первые 2 цифры: "91" из "9124"
        
        # Маппинг кодов на года
        year_map = {
            "91": 2021, "92": 2022, "93": 2023, "94": 2024,
            "95": 2025, "96": 2026, "97": 2027, "98": 2028,
            "99": 2029, "00": 2030, "01": 2031, "02": 2032,
        }
        
        return year_map.get(year_code)
    
    @staticmethod
    def determine_group_semester(
        group_name: str,
        semester_column: Optional[str],
        contingent_info: Optional[str],
        academic_year: str,
        is_autumn_semester: bool = True
    ) -> int:
        """
        Определяет семестр обучения группы (1-12).
        
        Args:
            group_name: "Б9124-01.03.02сп"
            semester_column: "Третий семестр" или None
            contingent_info: "...2 курс, гр Б9124..." или None
            academic_year: "2025/2026"
            is_autumn_semester: True для осеннего, False для весеннего
            
        Returns:
            Номер семестра обучения (1-12)
        """
        # 1. Попытка извлечь из колонки "Семестр"
        if semester_column and isinstance(semester_column, str):
            semester_num = ExcelParser._extract_semester_number(semester_column)
            if semester_num:
                return max(1, min(semester_num, 12))  # Ограничиваем 1-12
        
        # 2. Попытка вычислить из года поступления
        admission_year = ExcelParser.extract_admission_year(group_name)
        if admission_year:
            # Вычисляем текущий год из academic_year
            try:
                current_year = int(academic_year.split('/')[0])
            except (ValueError, AttributeError):
                # Если не удалось распарсить, используем текущий год
                from datetime import datetime
                current_year = datetime.now().year
            
            # Вычисляем курс
            course = (current_year - admission_year) + 1
            
            # Вычисляем семестр
            if is_autumn_semester:
                semester = course * 2 - 1  # Нечетный семестр (1, 3, 5, 7, 9, 11)
            else:
                semester = course * 2      # Четный семестр (2, 4, 6, 8, 10, 12)
            
            return max(1, min(semester, 12))  # Ограничиваем 1-12
        
        # 3. Попробовать извлечь курс из контингента
        if contingent_info:
            course_match = re.search(r'(\d+)\s*курс', contingent_info, re.IGNORECASE)
            if course_match:
                course = int(course_match.group(1))
                if is_autumn_semester:
                    semester = course * 2 - 1
                else:
                    semester = course * 2
                return max(1, min(semester, 12))
        
        # По умолчанию 1 семестр
        logger.warning(
            f"Could not determine semester for group {group_name}, "
            f"using default semester 1"
        )
        return 1
    
    @staticmethod
    def parse_multiple_files(
        files_data: List[Tuple[bytes, str]],
        academic_year: str
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Парсит множество Excel файлов от разных кафедр.
        
        Args:
            files_data: Список кортежей (байты файла, имя файла)
            academic_year: Учебный год "2024/2025"
            
        Returns:
            (объединенные_нагрузки, список_ошибок)
            
        Process:
            1. Парсит каждый файл отдельно
            2. Автоматически определяет семестр для каждой группы
            3. Объединяет все нагрузки
            4. Проверяет на дубликаты
            5. Возвращает единый список
        """
        all_loads = []
        all_errors = []
        
        # Определяем, какой сейчас семестр (осенний/весенний)
        # По умолчанию: месяцы 9-12 и 1 = осенний, 2-6 = весенний
        current_month = datetime.now().month
        is_autumn = current_month >= 9 or current_month == 1
        
        logger.info(
            f"Parsing {len(files_data)} files for academic year {academic_year} "
            f"(is_autumn={is_autumn})"
        )
        
        for file_data, filename in files_data:
            try:
                logger.info(f"Parsing file: {filename}")
                
                # Парсим один файл (используем существующий метод)
                loads, errors = ExcelParser.parse_course_loads(
                    file_data=file_data,
                    semester=1,  # Временное значение, будет перезаписано
                    academic_year=academic_year
                )
                
                if errors:
                    all_errors.extend([f"[{filename}] {err}" for err in errors])
                
                # Обновляем семестр для каждой нагрузки
                for load in loads:
                    group_name = load.get('group_name', '')
                    semester_col = load.get('semester_column')
                    contingent = load.get('contingent_info')
                    
                    # Автоматически определяем семестр
                    semester = ExcelParser.determine_group_semester(
                        group_name=group_name,
                        semester_column=semester_col,
                        contingent_info=contingent,
                        academic_year=academic_year,
                        is_autumn_semester=is_autumn
                    )
                    
                    load['semester'] = semester
                    load['source_file'] = filename  # Для отладки
                
                all_loads.extend(loads)
                logger.info(
                    f"Parsed {len(loads)} loads from {filename}, "
                    f"total loads: {len(all_loads)}"
                )
                
            except Exception as e:
                error_msg = f"[{filename}] Ошибка парсинга: {str(e)}"
                all_errors.append(error_msg)
                logger.error(error_msg, exc_info=True)
        
        # Проверка на дубликаты
        seen = set()
        deduplicated_loads = []
        
        for load in all_loads:
            # Уникальный ключ: группа + дисциплина + преподаватель + тип + семестр
            key = (
                load['group_name'],
                load['discipline_name'],
                load['teacher_name'],
                load['lesson_type'],
                load['semester']
            )
            
            if key not in seen:
                seen.add(key)
                deduplicated_loads.append(load)
            else:
                all_errors.append(
                    f"Дубликат: {load['discipline_name']} для {load['group_name']} "
                    f"({load['teacher_name']}, {load['lesson_type']}, семестр {load['semester']})"
                )
        
        logger.info(
            f"Total: {len(deduplicated_loads)} unique loads from {len(files_data)} files, "
            f"{len(all_errors)} errors/warnings"
        )
        
        return deduplicated_loads, all_errors

