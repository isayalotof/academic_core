"""
⭐ Excel Parser для реального формата учебной нагрузки
Гибкий парсер с автоматическим определением формата (20 колонок)
"""
import openpyxl
from typing import Dict, List, Any, Optional
import logging
import re
import io

logger = logging.getLogger(__name__)


class CourseLoadExcelParser:
    """Парсер Excel с учебной нагрузкой для реального формата"""
    
    # КРИТИЧНО: 16 недель в семестре!
    WEEKS_IN_SEMESTER = 16
    HOURS_PER_LESSON = 1.5  # астрономических часа (90 минут)
    
    # Маппинг колонок (гибкий поиск по названиям)
    COLUMN_PATTERNS = {
        'discipline': ['дисциплина', 'название дисциплины'],
        'load_type': ['нагрузка', 'вид занятия', 'тип занятия'],
        'block': ['блок', 'код дисциплины', 'код'],
        'semester': ['семестр', 'сем'],
        'hours': ['часы', 'часов', 'часов в семестр', 'количество часов'],
        'teacher': ['преподаватель', 'фио преподавателя', 'фио'],
        'group': ['группа', 'название группы'],
        'students_count': ['количество контингента', 'количество', 'контингент'],
        'department': ['кафедра', 'кафедр'],
    }
    
    # Маппинг типов занятий
    LESSON_TYPE_MAPPING = {
        'лекционные занятия': 'Лекция',
        'лекции': 'Лекция',
        'практические занятия': 'Практика',
        'практика': 'Практика',
        'семинарские занятия': 'Практика',
        'лабораторные работы': 'Лабораторная',
        'лабораторные занятия': 'Лабораторная',
        'лаборатор': 'Лабораторная',
        'консультации': 'Консультация',
        'индивидуальные консультации': 'Консультация',
        'консультации перед экзаменом': 'Консультация',
        'индивидуальные консультации по лабораторным': 'Консультация',
    }
    
    # Маппинг семестров (текст → номер)
    SEMESTER_WORDS = {
        'первый': 1, 'второй': 2, 'третий': 3, 
        'четвертый': 4, 'четвёртый': 4,
        'пятый': 5, 'шестой': 6, 'седьмой': 7, 'восьмой': 8,
    }
    
    def parse(self, file_data: bytes, filename: str,
              semester: Optional[int] = None,
              academic_year: Optional[str] = None) -> Dict[str, Any]:
        """
        Парсинг Excel файла
        
        Returns:
            {
                'success': bool,
                'data': List[Dict],
                'errors': List[str],
                'warnings': List[str],
                'total_rows': int,
                'stats': Dict
            }
        """
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, read_only=True)
            ws = wb.active
            
            # Найти заголовки
            header_row_idx = self._find_header_row(ws)
            if not header_row_idx:
                return {
                    'success': False,
                    'errors': ['Не найдены заголовки таблицы'],
                    'data': [],
                    'warnings': [],
                    'total_rows': 0,
                    'stats': {}
                }
            
            # Маппинг колонок
            column_mapping = self._map_columns(ws, header_row_idx)
            
            # Валидация
            missing = self._validate_required_columns(column_mapping)
            if missing:
                return {
                    'success': False,
                    'errors': [f'Отсутствуют колонки: {", ".join(missing)}'],
                    'data': [],
                    'warnings': [],
                    'total_rows': 0,
                    'stats': {}
                }
            
            # Парсинг строк
            data = []
            errors = []
            warnings = []
            stats = {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'by_lesson_type': {},
                'by_teacher': {},
                'vacancies': 0,
            }
            
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                stats['total_processed'] += 1
                
                try:
                    row_data = self._parse_row(
                        ws, row_idx, column_mapping, 
                        semester, academic_year
                    )
                    
                    if row_data is None:
                        continue
                    
                    if row_data.get('_skip'):
                        continue
                    
                    data.append(row_data)
                    stats['successful'] += 1
                    
                    # Статистика
                    lesson_type = row_data.get('lesson_type', 'Unknown')
                    stats['by_lesson_type'][lesson_type] = \
                        stats['by_lesson_type'].get(lesson_type, 0) + 1
                    
                    teacher_name = row_data.get('teacher_name', '')
                    if teacher_name and teacher_name.lower() not in ['вакансия', 'vacancy']:
                        stats['by_teacher'][teacher_name] = \
                            stats['by_teacher'].get(teacher_name, 0) + 1
                    else:
                        stats['vacancies'] += 1
                    
                except Exception as e:
                    stats['failed'] += 1
                    error_msg = f"Строка {row_idx}: {str(e)}"
                    errors.append(error_msg)
                    logger.warning(error_msg)
            
            logger.info(
                f"Parsed {stats['successful']} rows, "
                f"{stats['failed']} errors, "
                f"{stats['vacancies']} vacancies"
            )
            
            return {
                'success': len(data) > 0,
                'data': data,
                'errors': errors,
                'warnings': warnings,
                'total_rows': len(data),
                'stats': stats
            }
            
        except Exception as e:
            logger.error(f"Critical error: {e}", exc_info=True)
            return {
                'success': False,
                'errors': [f"Критическая ошибка: {str(e)}"],
                'data': [],
                'warnings': [],
                'total_rows': 0,
                'stats': {}
            }
        finally:
            if 'wb' in locals():
                wb.close()
    
    def _find_header_row(self, ws) -> Optional[int]:
        """Найти строку с заголовками"""
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_values = [
                str(cell.value).lower() if cell.value else '' 
                for cell in ws[row_idx]
            ]
            
            has_discipline = any('дисциплина' in val for val in row_values)
            has_teacher = any('преподаватель' in val for val in row_values)
            
            if has_discipline and has_teacher:
                return row_idx
        
        return None
    
    def _map_columns(self, ws, header_row_idx: int) -> Dict[str, int]:
        """Создать маппинг колонок"""
        mapping = {}
        header_row = ws[header_row_idx]
        
        for col_idx, cell in enumerate(header_row):
            if not cell.value:
                continue
            
            header_lower = str(cell.value).lower().strip()
            
            for field, patterns in self.COLUMN_PATTERNS.items():
                if any(pattern in header_lower for pattern in patterns):
                    mapping[field] = col_idx
                    break
        
        return mapping
    
    def _validate_required_columns(self, mapping: Dict) -> bool:
        """Проверить обязательные колонки"""
        required = ['discipline', 'hours', 'teacher', 'group']
        return all(f in mapping for f in required)
    
    def _parse_row(self, ws, row_idx: int, mapping: Dict,
                   filter_semester: Optional[int],
                   academic_year: Optional[str]) -> Optional[Dict]:
        """Парсинг одной строки"""
        
        row = ws[row_idx]
        
        # Извлечь значения
        discipline_name = self._get_cell_value(row, mapping.get('discipline'))
        teacher_name = self._get_cell_value(row, mapping.get('teacher'))
        group_name = self._get_cell_value(row, mapping.get('group'))
        hours_raw = self._get_cell_value(row, mapping.get('hours'))
        
        # Проверка пустой строки
        if not all([discipline_name, teacher_name, group_name, hours_raw]):
            return None
        
        # Парсинг часов
        try:
            hours = float(hours_raw)
        except (ValueError, TypeError):
            raise ValueError(f"Некорректные часы: '{hours_raw}'")
        
        if hours <= 0:
            return None
        
        # Тип занятия
        load_type_raw = self._get_cell_value(row, mapping.get('load_type'))
        lesson_type = self._normalize_lesson_type(load_type_raw) \
            if load_type_raw else 'Практика'
        
        # Семестр
        semester_raw = self._get_cell_value(row, mapping.get('semester'))
        semester = self._extract_semester(semester_raw) \
            if semester_raw else 1
        
        # Фильтр по семестру
        if filter_semester and semester != filter_semester:
            return {'_skip': True}
        
        # Преподаватель
        teacher_info = self._parse_teacher_name(teacher_name)
        
        # Пропустить вакансии
        if teacher_info['is_vacancy']:
            return {'_skip': True}
        
        # Группа
        group_info = self._parse_group_name(group_name)
        
        # Количество студентов
        students_count_raw = self._get_cell_value(
            row, mapping.get('students_count')
        )
        try:
            students_count = int(float(students_count_raw)) \
                if students_count_raw else None
        except (ValueError, TypeError):
            students_count = None
        
        # Код дисциплины
        discipline_code = self._get_cell_value(row, mapping.get('block'))
        
        # Кафедра
        department = self._get_cell_value(row, mapping.get('department'))
        
        # КРИТИЧНО: Рассчитать пары в неделю
        # Формула: часы / 24 (где 24 = 1.5 часа * 16 недель)
        lessons_per_week = self._calculate_lessons_per_week(hours)
        
        # Собрать результат
        return {
            # Дисциплина
            'discipline_name': discipline_name.strip(),
            'discipline_code': discipline_code.strip() 
                if discipline_code else None,
            'department': department.strip() if department else None,
            
            # Преподаватель
            'teacher_name': teacher_info['full_name'],
            'teacher_first_name': teacher_info.get('first_name'),
            'teacher_last_name': teacher_info.get('last_name'),
            'teacher_middle_name': teacher_info.get('middle_name'),
            'is_vacancy': False,
            
            # Группа
            'group_name': group_info['name'],
            'group_short_name': group_info['short_name'],
            'group_year': group_info['year'],
            'group_level': group_info['level'],
            'group_program_code': group_info.get('program_code'),
            'group_specialization': group_info.get('specialization'),
            'group_subgroup': group_info.get('subgroup'),
            'students_count': students_count,
            
            # Нагрузка
            'lesson_type': lesson_type,
            'hours_per_semester': hours,
            'lessons_per_week': lessons_per_week,  # ⭐ Рассчитано
            'weeks_count': self.WEEKS_IN_SEMESTER,
            
            # Семестр
            'semester': semester,
            'academic_year': academic_year or \
                self._guess_academic_year(group_info['year'], semester),
            
            # Мета
            'source_row': row_idx,
        }
    
    def _get_cell_value(self, row, col_idx: Optional[int]) -> Optional[str]:
        """Получить значение ячейки"""
        if col_idx is None:
            return None
        if col_idx >= len(row):
            return None
        cell = row[col_idx]
        value = cell.value
        return str(value).strip() if value is not None else None
    
    def _normalize_lesson_type(self, raw_type: str) -> str:
        """Нормализация типа занятия"""
        raw_lower = raw_type.lower().strip()
        
        for key, value in self.LESSON_TYPE_MAPPING.items():
            if key in raw_lower:
                return value
        
        return 'Практика'
    
    def _extract_semester(self, semester_text: str) -> int:
        """Извлечь номер семестра"""
        text_lower = semester_text.lower().strip()
        
        # Поиск по словам
        for word, number in self.SEMESTER_WORDS.items():
            if word in text_lower:
                return number
        
        # Поиск цифр
        digits = re.findall(r'\d+', text_lower)
        if digits:
            return int(digits[0])
        
        return 1
    
    def _parse_teacher_name(self, raw_name: str) -> Dict[str, Any]:
        """Парсинг ФИО преподавателя"""
        if not raw_name or not raw_name.strip():
            return {
                'full_name': 'Вакансия',
                'is_vacancy': True,
            }
        
        raw_name = raw_name.strip()
        
        if 'вакансия' in raw_name.lower():
            return {
                'full_name': 'Вакансия',
                'is_vacancy': True,
            }
        
        # Разбить ФИО
        parts = raw_name.split()
        
        if len(parts) >= 3:
            last_name, first_name, middle_name = parts[0], parts[1], parts[2]
        elif len(parts) == 2:
            last_name, first_name = parts[0], parts[1]
            middle_name = None
        else:
            last_name = raw_name
            first_name = None
            middle_name = None
        
        return {
            'full_name': raw_name,
            'is_vacancy': False,
            'last_name': last_name,
            'first_name': first_name,
            'middle_name': middle_name
        }
    
    def _parse_group_name(self, raw_group: str) -> Dict[str, Any]:
        """Парсинг названия группы"""
        if not raw_group or not raw_group.strip():
            raise ValueError("Пустое название группы")
        
        raw_group = raw_group.strip()
        
        # Уровень (Б=бакалавр, М=магистр, А=аспирантура)
        level_map = {'Б': 'bachelor', 'М': 'master', 'А': 'postgraduate'}
        level = level_map.get(raw_group[0], 'bachelor')
        
        # Год (Б9124 → 91 → 2024)
        year_match = re.search(r'[БМА](\d{2})(\d{2})', raw_group)
        if year_match:
            year_short = year_match.group(1)
            year = 2000 + int(year_short)
        else:
            year = None
        
        # Короткое название (Б9124)
        short_match = re.match(r'([БМА]\d{4})', raw_group)
        short_name = short_match.group(1) if short_match else raw_group
        
        # Подгруппа (1) из "Б9124-01.03.02сп(1)"
        subgroup_match = re.search(r'\((\d+)\)', raw_group)
        subgroup = int(subgroup_match.group(1)) if subgroup_match else None
        
        # Код направления (01.03.02)
        program_match = re.search(r'(\d{2}\.\d{2}\.\d{2})', raw_group)
        program_code = program_match.group(1) if program_match else None
        
        # Специализация (сп)
        specialization = None
        if program_code:
            search_str = re.sub(r'\(\d+\)', '', raw_group)
            spec_match = re.search(f'{re.escape(program_code)}([а-яА-Я]+)', search_str)
            if spec_match:
                specialization = spec_match.group(1)
        
        return {
            'name': raw_group,
            'short_name': short_name,
            'year': year,
            'level': level,
            'subgroup': subgroup,
            'program_code': program_code,
            'specialization': specialization
        }
    
    def _calculate_lessons_per_week(self, hours: float) -> int:
        """
        КРИТИЧНО: Рассчитать пары в неделю
        
        Формула: часы / (1.5 * 16) = часы / 24
        
        Примеры:
        - 72 часа → 3 пары
        - 36 часов → 2 пары (округление от 1.5)
        - 18 часов → 1 пара
        - 2 часа → 1 пара (минимум)
        """
        lessons = hours / (self.HOURS_PER_LESSON * self.WEEKS_IN_SEMESTER)
        
        # Округление
        lessons_rounded = round(lessons)
        
        # Минимум 1 пара если часы > 0
        if hours > 0 and lessons_rounded < 1:
            lessons_rounded = 1
        
        return lessons_rounded
    
    def _guess_academic_year(self, group_year: Optional[int], 
                            semester: int) -> str:
        """Угадать учебный год"""
        if not group_year:
            return "2024/2025"
        
        # Семестры 1-2 → 1 курс, 3-4 → 2 курс
        course = (semester + 1) // 2
        study_year = group_year + course - 1
        
        return f"{study_year}/{study_year + 1}"


def parse_course_loads_excel(file_data: bytes, filename: str,
                             semester: Optional[int] = None,
                             academic_year: Optional[str] = None) -> Dict[str, Any]:
    """
    Функция-обёртка для парсинга Excel
    
    Args:
        file_data: Байты файла
        filename: Имя файла
        semester: Фильтр по семестру (опционально)
        academic_year: Учебный год (опционально)
    
    Returns:
        Результат парсинга
    """
    parser = CourseLoadExcelParser()
    return parser.parse(file_data, filename, semester, academic_year)
