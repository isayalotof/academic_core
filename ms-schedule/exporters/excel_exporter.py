"""
Excel exporter for schedule
Экспорт расписания в Excel формат с красивым форматированием
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict
import io
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Экспорт расписания в Excel"""
    
    # Русские названия дней недели
    DAYS_RU = {
        1: "Понедельник",
        2: "Вторник",
        3: "Среда",
        4: "Четверг",
        5: "Пятница",
        6: "Суббота"
    }
    
    # Временные слоты
    TIME_SLOTS = {
        1: "09:00 - 10:30",
        2: "10:45 - 12:15",
        3: "13:00 - 14:30",
        4: "14:45 - 16:15",
        5: "16:30 - 18:00",
        6: "18:15 - 19:45"
    }
    
    # Цвета для разных типов занятий
    LESSON_COLORS = {
        "лекция": "CCE5FF",        # Голубой
        "практика": "E5FFCC",      # Зеленый
        "лабораторная": "FFE5CC",  # Оранжевый
        "default": "F0F0F0"        # Серый
    }
    
    def export_group_schedule(
        self, 
        lessons: List[Dict], 
        group_name: str,
        semester: int,
        academic_year: str
    ) -> bytes:
        """
        Экспорт расписания группы в Excel
        
        Args:
            lessons: Список занятий
            group_name: Название группы
            semester: Номер семестра
            academic_year: Учебный год
            
        Returns:
            Excel file as bytes
        """
        logger.info(f"Exporting schedule for group {group_name}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{group_name}"[:31]  # Excel limit
        
        # Заголовок
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'Расписание группы {group_name} ({semester} семестр, {academic_year})'
        title_cell.font = Font(size=16, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
        
        ws.row_dimensions[1].height = 30
        
        # Шапка таблицы
        headers = ['День', 'Время', 'Дисциплина', 'Преподаватель', 'Тип', 'Аудитория', 'Недели']
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
        
        ws.row_dimensions[3].height = 25
        
        # Данные
        current_row = 4
        last_day = None
        
        for lesson in sorted(lessons, key=lambda x: (x['day_of_week'], x['time_slot'])):
            # Разделитель между днями
            if last_day and last_day != lesson['day_of_week']:
                ws.row_dimensions[current_row].height = 3
                current_row += 1
            
            # День недели
            cell = ws.cell(row=current_row, column=1)
            cell.value = self.DAYS_RU[lesson['day_of_week']]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
            
            # Время
            cell = ws.cell(row=current_row, column=2)
            cell.value = self.TIME_SLOTS.get(lesson['time_slot'], f"Пара {lesson['time_slot']}")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
            
            # Дисциплина
            cell = ws.cell(row=current_row, column=3)
            cell.value = lesson['discipline_name']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = self._get_border()
            
            # Преподаватель
            cell = ws.cell(row=current_row, column=4)
            cell.value = lesson['teacher_name']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = self._get_border()
            
            # Тип занятия
            cell = ws.cell(row=current_row, column=5)
            cell.value = lesson['lesson_type']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
            
            # Цвет по типу
            lesson_type_lower = lesson['lesson_type'].lower()
            color = self.LESSON_COLORS.get(
                next((key for key in self.LESSON_COLORS if key in lesson_type_lower), 'default'),
                self.LESSON_COLORS['default']
            )
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            
            # Аудитория
            cell = ws.cell(row=current_row, column=6)
            classroom = lesson.get('classroom_name', '-')
            building = lesson.get('building_name', '')
            cell.value = f"{classroom} ({building})" if building and classroom != '-' else classroom
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
            
            # Недели (числитель/знаменатель)
            cell = ws.cell(row=current_row, column=7)
            week_type = lesson.get('week_type', 'both')
            week_text = {
                'odd': 'Числ.',
                'even': 'Знам.',
                'both': 'Обе'
            }.get(week_type, week_type)
            cell.value = week_text
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
            
            ws.row_dimensions[current_row].height = 35
            
            last_day = lesson['day_of_week']
            current_row += 1
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 15  # День
        ws.column_dimensions['B'].width = 15  # Время
        ws.column_dimensions['C'].width = 35  # Дисциплина
        ws.column_dimensions['D'].width = 25  # Преподаватель
        ws.column_dimensions['E'].width = 12  # Тип
        ws.column_dimensions['F'].width = 15  # Аудитория
        ws.column_dimensions['G'].width = 10  # Недели
        
        # Сохранить в bytes
        output = io.BytesIO()
        wb.save(output)
        
        logger.info(f"Excel export completed: {len(lessons)} lessons")
        return output.getvalue()
    
    def export_teacher_schedule(
        self,
        lessons: List[Dict],
        teacher_name: str,
        semester: int,
        academic_year: str
    ) -> bytes:
        """Экспорт расписания преподавателя"""
        logger.info(f"Exporting schedule for teacher {teacher_name}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"
        
        # Заголовок
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'Расписание преподавателя {teacher_name} ({semester} семестр, {academic_year})'
        title_cell.font = Font(size=16, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
        ws.row_dimensions[1].height = 30
        
        # Шапка
        headers = ['День', 'Время', 'Дисциплина', 'Группа', 'Тип', 'Аудитория', 'Недели']
        header_fill = PatternFill(start_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
        
        # Данные (аналогично группе, но с группой вместо преподавателя)
        current_row = 4
        for lesson in sorted(lessons, key=lambda x: (x['day_of_week'], x['time_slot'])):
            ws.cell(row=current_row, column=1).value = self.DAYS_RU[lesson['day_of_week']]
            ws.cell(row=current_row, column=2).value = self.TIME_SLOTS.get(lesson['time_slot'])
            ws.cell(row=current_row, column=3).value = lesson['discipline_name']
            ws.cell(row=current_row, column=4).value = lesson['group_name']
            ws.cell(row=current_row, column=5).value = lesson['lesson_type']
            
            classroom = lesson.get('classroom_name', '-')
            building = lesson.get('building_name', '')
            ws.cell(row=current_row, column=6).value = f"{classroom} ({building})" if building else classroom
            
            week_type = lesson.get('week_type', 'both')
            ws.cell(row=current_row, column=7).value = {
                'odd': 'Числ.', 'even': 'Знам.', 'both': 'Обе'
            }.get(week_type, week_type)
            
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self._get_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # Ширина колонок
        for col, width in zip(range(1, 8), [15, 15, 35, 20, 12, 15, 10]):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _get_border(self) -> Border:
        """Создать стандартную границу"""
        thin_side = Side(style='thin', color="000000")
        return Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )


# Global instance
excel_exporter = ExcelExporter()

